import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

from modules.reporting import (
    aggregate_ecl_by_dimension,
    aggregate_ecl_by_stage,
    build_excel_export_bytes,
    build_dashboard_metrics,
    build_review_flags,
)


def test_dashboard_metrics_include_stage_shares_and_review_count():
    portfolio = pd.DataFrame(
        [
            {"loan_id": "LN-1", "stage": "Stage 1", "ead": 100, "ecl": 1, "review_required": False},
            {"loan_id": "LN-2", "stage": "Stage 2", "ead": 200, "ecl": 20, "review_required": True},
            {"loan_id": "LN-3", "stage": "Stage 3", "ead": 300, "ecl": 120, "review_required": False},
        ]
    )
    findings = pd.DataFrame([{"loan_id": "LN-2", "check_code": "MISSING_PD", "description": "PD manquante"}])

    metrics = build_dashboard_metrics(portfolio, findings)

    assert metrics["total_ead"] == 600
    assert metrics["total_ecl"] == 141
    assert metrics["stage_2_share"] == 1 / 3
    assert metrics["stage_3_share"] == 1 / 3
    assert metrics["data_quality_issue_count"] == 1
    assert metrics["review_required_count"] == 1


def test_review_flags_detect_data_quality_and_threshold_cases():
    portfolio = pd.DataFrame(
        [
            {
                "loan_id": "LN-1",
                "days_past_due": 28,
                "ecl": 10,
            },
            {
                "loan_id": "LN-2",
                "days_past_due": 0,
                "ecl": 1000,
            },
        ]
    )
    findings = pd.DataFrame(
        [{"loan_id": "LN-1", "check_code": "MISSING_RATING", "description": "Rating manquant"}]
    )

    result = build_review_flags(portfolio, findings)

    assert bool(result.loc[0, "review_required"]) is True
    assert bool(result.loc[0, "rating_missing"]) is True
    assert bool(result.loc[0, "dpd_near_30"]) is True
    assert bool(result.loc[1, "high_ecl_contribution"]) is True


def test_aggregates_ecl_by_stage_product_and_sector():
    portfolio = pd.DataFrame(
        [
            {"loan_id": "LN-1", "stage": "Stage 1", "product_type": "SME", "sector": "Retail", "ead": 100, "ecl": 2},
            {"loan_id": "LN-2", "stage": "Stage 1", "product_type": "SME", "sector": "Retail", "ead": 200, "ecl": 3},
            {"loan_id": "LN-3", "stage": "Stage 2", "product_type": "Mortgage", "sector": "Real estate", "ead": 300, "ecl": 30},
        ]
    )

    by_stage = aggregate_ecl_by_stage(portfolio)
    by_product = aggregate_ecl_by_dimension(portfolio, "product_type")
    by_sector = aggregate_ecl_by_dimension(portfolio, "sector")

    assert by_stage.loc[by_stage["stage"].eq("Stage 1"), "ecl"].iloc[0] == 5
    assert by_stage.loc[by_stage["stage"].eq("Stage 2"), "coverage_ratio"].iloc[0] == 0.1
    assert by_product.loc[by_product["product_type"].eq("SME"), "ead"].iloc[0] == 300
    assert by_sector.loc[by_sector["sector"].eq("Retail"), "exposure_count"].iloc[0] == 2


def test_excel_export_accepts_audit_trail_and_committee_summary():
    payload = build_excel_export_bytes(
        pd.DataFrame({"loan_id": ["LN-1"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "check_code": ["MISSING_PD"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "stage": ["Stage 1"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "ecl": [10]}),
        pd.DataFrame({"metric": ["ECL"], "value": [10]}),
        {"run_summary": pd.DataFrame({"item": ["run"], "value": ["RUN-1"]})},
        detailed_audit_trail={"run_summary": pd.DataFrame({"field": ["run_id"], "value": ["RUN-1"]})},
        committee_summary="# Note de synthese",
    )

    assert len(payload) > 0


def test_excel_export_includes_v06_business_demo_sheets():
    payload = build_excel_export_bytes(
        pd.DataFrame({"loan_id": ["LN-1"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "check_code": ["MISSING_PD"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "stage": ["Stage 1"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "ecl": [10]}),
        pd.DataFrame({"metric": ["Business consistency score"], "value": [0.99]}),
        {"run_summary": pd.DataFrame({"item": ["run"], "value": ["RUN-1"]})},
        business_consistency=pd.DataFrame({"loan_id": ["LN-1"], "severity": ["Warning"]}),
        demo_storyline=pd.DataFrame({"step": [1], "title": ["Portfolio"]}),
        client_discussion_points=pd.DataFrame({"discussion_point": ["Question client"]}),
    )

    workbook = load_workbook(BytesIO(payload), read_only=True)

    assert "Business Consistency" in workbook.sheetnames
    assert "Demo Storyline" in workbook.sheetnames
    assert "Client Discussion Points" in workbook.sheetnames


def test_excel_export_includes_v2_risk_parameter_sheets():
    payload = build_excel_export_bytes(
        pd.DataFrame({"loan_id": ["LN-1"]}),
        pd.DataFrame(columns=["loan_id", "check_code"]),
        pd.DataFrame({"loan_id": ["LN-1"], "stage": ["Stage 1"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "ecl": [10]}),
        pd.DataFrame({"metric": ["ECL"], "value": [10]}),
        {"run_summary": pd.DataFrame({"item": ["run"], "value": ["RUN-1"]})},
        risk_parameters=pd.DataFrame(
            {"loan_id": ["LN-1"], "pd_12m": [0.02], "pd_lifetime": [0.05]}
        ),
        lifetime_pd_curve=pd.DataFrame(
            {"stage": ["Stage 1"], "year": [1], "cumulative_pd": [0.02]}
        ),
        lgd_parameters=pd.DataFrame(
            {"loan_id": ["LN-1"], "lgd": [0.40], "collateral_haircut": [0.20]}
        ),
        lgd_sensitivity=pd.DataFrame(
            {"scenario": ["Baseline"], "lgd": [0.40]}
        ),
        ead_parameters=pd.DataFrame(
            {
                "loan_id": ["LN-1"],
                "ead_accounting": [100.0],
                "undrawn_commitment": [20.0],
                "ccf_adjusted": [0.50],
            }
        ),
        ead_curve=pd.DataFrame(
            {
                "product_type": ["SME term loan"],
                "year": [1],
                "ead_projected": [90.0],
            }
        ),
    )

    workbook = load_workbook(BytesIO(payload), read_only=True)

    assert "Risk Parameters" in workbook.sheetnames
    assert "Lifetime PD Curve" in workbook.sheetnames
    assert "LGD Parameters" in workbook.sheetnames
    assert "LGD Sensitivity" in workbook.sheetnames
    assert "EAD Parameters" in workbook.sheetnames
    assert "EAD Curve" in workbook.sheetnames
