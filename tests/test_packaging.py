from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

from modules.committee_summary import generate_committee_summary
from modules.demo_config import DEMO_DISCLAIMER_FR
from modules.reporting import build_excel_export_bytes


def test_main_modules_import_cleanly():
    import modules.audit_trail  # noqa: F401
    import modules.business_checks  # noqa: F401
    import modules.calculation_utils  # noqa: F401
    import modules.committee_summary  # noqa: F401
    import modules.data_quality  # noqa: F401
    import modules.ecl_calculator  # noqa: F401
    import modules.lgd_engine  # noqa: F401
    import modules.overlay_engine  # noqa: F401
    import modules.reporting  # noqa: F401
    import modules.risk_parameters  # noqa: F401
    import modules.sample_data  # noqa: F401
    import modules.scenario_engine  # noqa: F401
    import modules.staging_engine  # noqa: F401
    import ui.branding  # noqa: F401
    import ui.components  # noqa: F401
    import ui.theme  # noqa: F401


def test_packaging_docs_exist():
    required_docs = [
        "docs/USER_GUIDE.md",
        "docs/DEMO_SCRIPT.md",
        "docs/METHODOLOGY_NOTES.md",
        "docs/ROADMAP.md",
    ]
    for doc_path in required_docs:
        assert Path(doc_path).exists()


def test_public_demo_does_not_offer_external_data_import():
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert "file_uploader" not in app_source
    assert "read_uploaded_file" not in app_source
    assert "Charger un fichier" not in app_source


def test_streamlit_shell_is_split_into_ui_modules():
    app_source = Path("app.py").read_text(encoding="utf-8")

    assert "from ui.theme import apply_auria_theme" in app_source
    assert "from ui.branding import render_brand_header" in app_source
    assert "from ui.components import" in app_source
    assert Path("ui/theme.py").exists()
    assert Path("ui/branding.py").exists()
    assert Path("ui/components.py").exists()


def test_disclaimer_is_present_in_committee_summary_and_excel_export():
    note = generate_committee_summary(
        "RUN-1",
        {"total_ead": 1000, "total_ecl": 50, "coverage_ratio": 0.05, "exposure_count": 2, "data_quality_issue_count": 0},
        {"ecl_weighted": 55, "weighted_impact_amount": 5, "weighted_impact_pct": 0.10, "downside_impact_amount": 10, "downside_impact_pct": 0.20},
        {"ecl_before_overlay": 55, "ecl_after_overlay": 60, "total_overlay_amount": 5, "overlay_variation_pct": 0.09, "top_overlay_contributor": "None"},
        pd.DataFrame({"stage": ["Stage 1"], "exposure_count": [2], "ead": [1000], "ecl": [50]}),
        pd.DataFrame({"product_type": ["SME"], "ecl": [50]}),
        pd.DataFrame({"sector": ["Energy"], "ecl": [50]}),
        pd.DataFrame({"stage": ["Stage 1"], "count": [2]}),
        pd.DataFrame({"scenario": ["Baseline"], "weight": [1.0]}),
        pd.DataFrame({"scenario": ["Baseline"], "weight": [1.0], "ecl": [55]}),
        pd.DataFrame({"overlay_name": ["None"], "overlay_amount": [0], "justification": ["N/A"]}),
        pd.DataFrame(columns=["check_code", "issue_count"]),
        0,
        pd.DataFrame({"loan_id": ["LN-1"], "ecl": [50], "stage": ["Stage 1"]}),
        ["Message cle"],
    )
    assert DEMO_DISCLAIMER_FR in note

    payload = build_excel_export_bytes(
        pd.DataFrame({"loan_id": ["LN-1"]}),
        pd.DataFrame(columns=["loan_id", "check_code"]),
        pd.DataFrame({"loan_id": ["LN-1"], "stage": ["Stage 1"]}),
        pd.DataFrame({"loan_id": ["LN-1"], "ecl": [10]}),
        pd.DataFrame({"metric": ["ECL"], "value": [10]}),
        {"run_summary": pd.DataFrame({"item": ["run"], "value": ["RUN-1"]})},
        committee_summary=note,
    )
    workbook = load_workbook(BytesIO(payload), read_only=True)
    assert "Disclaimer" in workbook.sheetnames
